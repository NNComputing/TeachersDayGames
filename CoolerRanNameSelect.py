'''
Don't use this cell
        - Aakarsh
Code uses cool functions but doesn't seem so random for some reason
'''

import openpyxl as xl
import random

x = 1
not_playing = ['aakarsh', 'sayem', 'didum', 'navin']

# Load the Excel workbook
workbook = xl.load_workbook('2412_NAMES.xlsx')  # Replace '2412_NAMES.xlsx' with your Excel file name
sheet = workbook.active

# Get the number of rows in the sheet
num_rows = sheet.max_row

# Read names from the sheet and store in a list
names = []
for row in sheet.iter_rows(min_row=2, max_row=num_rows, min_col=1, max_col=1):
    name = row[0].value
    names.append(name)

# Select a random name from the list
def generate_unique_names(names, not_playing):
    while True:
        q_name = random.choice(names).lower()
        l_name = random.choice(names).lower()

        if q_name in not_playing or l_name in not_playing:
            continue  # Retry if names are in the not_playing list

        if q_name == l_name:
            continue  # Retry if q_name is the same as l_name

        return q_name, l_name

q_name, l_name = generate_unique_names(names, not_playing)
print(f"{q_name.capitalize()} has to ask {l_name.capitalize()} a question")

#### BY NAVIN AMAL ANAND and KARTHIKEYAN AAKARSHA KANNAN ASRJC CLASS 24/12 ####
