'''
USE THIS CELL
    - Aakarsh
This version takes into account those who are not playing
'''
import openpyxl as xl
import random

x=1
not_playing=['aakarsh', 'sayem', 'didum', 'navin']#add those who are not here with lowercase
# Load the Excel workbook
workbook = xl.load_workbook('2412_NAMES.xlsx')  # Replace '2412_NAMES.xlsx' with your Excel file name

# Select the appropriate sheet
sheet = workbook.active  # You can use sheet = workbook['SheetName'] to select a specific sheet

# Get the number of rows in the sheet
num_rows = sheet.max_row

# Read names from the sheet and store in a list
names = []
for row in sheet.iter_rows(min_row=2, max_row=num_rows, min_col=1, max_col=1):
    name = row[0].value
    names.append(name)

# Select a random name from the list
while x==1:
  q_name = random.choice(names).lower()
  l_name = random.choice(names).lower()
  for i in range(len(not_playing)):
    if q_name == not_playing[i] or l_name == not_playing[i]:
      q_name = random.choice(names).lower()
      l_name = random.choice(names).lower()
  if q_name == l_name:
      q_name = random.choice(names).lower()
      l_name = random.choice(names).lower()
  else:
    x=2


print(f"{q_name.capitalize()} has to ask {l_name.capitalize()} a question")


#### BY NAVIN AMAL ANAND and KARTHIKEYAN AAKARSHA KANNAN ASRJC CLASS 24/12 ####
