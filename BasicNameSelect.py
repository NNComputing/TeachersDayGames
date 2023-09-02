'''
Basic version
'''

import openpyxl as xl
import random

x=1
# Load the Excel workbook
workbook = xl.load_workbook('2412_NAMES.xlsx')  # Replace '2412.xlsx' with your Excel file name

# Select the appropriate sheet
sheet = workbook.active  # You can use sheet = workbook['SheetName'] to select a specific sheet

# Get the number of rows in the sheet
num_rows = sheet.max_row

# Read names from the sheet and store in a list
names = []
for row in sheet.iter_rows(min_row=2, max_row=num_rows, min_col=1, max_col=1):
    name = row[0].value
    names.append(name)

# Select a random name from the list
while x==1:
  q_name = random.choice(names).capitalize()
  l_name = random.choice(names).capitalize()
  if q_name == l_name:
      q_name = random.choice(names).capitalize()
      l_name = random.choice(names).capitalize()
  else:
    x=2


print(f"{q_name} has to ask {l_name} a question")

#### BY NAVIN AMAL ANAND and KARTHIKEYAN AAKARSHA KANNAN ASRJC CLASS 24/12 ####
